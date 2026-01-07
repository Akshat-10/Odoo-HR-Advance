from odoo import models, api
from odoo.exceptions import UserError
from io import BytesIO
import base64
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import io
from PIL import Image as PILImage
from openpyxl.drawing.image import Image


class HrCustomFormDeptAttendanceReport(models.Model):
    _inherit = "hr.custom.form.dept_attendance"

    def action_print_department_attendance_excel(self):
        if not self.line_ids:
            raise UserError("No department lines found.")

        wb = Workbook()
        ws = wb.active
        ws.title = "Department Attendance"

        if self.env.company.logo:
            max_width, max_height = 100, 80  # adjust size here
            image_data = base64.b64decode(self.env.company.logo)
            image = PILImage.open(io.BytesIO(image_data))

            width, height = image.size
            aspect_ratio = width / height

            if width > max_width:
                width = max_width
                height = int(width / aspect_ratio)

            if height > max_height:
                height = max_height
                width = int(height * aspect_ratio)

            resized_image = image.resize((width, height), PILImage.LANCZOS)

            img_bytes = io.BytesIO()
            resized_image.save(img_bytes, format='PNG')
            img_bytes.seek(0)

            logo = Image(img_bytes)
            logo.anchor = "A1"
            ws.add_image(logo)

        # ================= STYLES =================
        header_fill = PatternFill("solid", fgColor="5B9BD5")
        sub_header_fill = PatternFill("solid", fgColor="9CDCF3")
        bold = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        thin = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # ======== 🔴 CHANGE START: vertical divider style ========
        thin_right = Side(style="thin")
        # ======== 🔴 CHANGE END ========

        ws.row_dimensions[1].height = 61
        ws.row_dimensions[2].height = 35

        ws.merge_cells("B1:B1")  # Company name only in B
        ws.merge_cells("C1:E1")  # Date in C-D-E


        ws["B1"] = self.company_id.name
        ws["B1"].font = Font(bold=True, size=14)
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["B1"].fill = header_fill


        ws["C1"] = f"DATE {self.form_date.strftime('%d/%m/%Y')}"
        ws["C1"].font = Font(bold=True)
        ws["C1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["C1"].fill = header_fill

        # ================= HEADERS =================
        headers = [
            "SR. NO.",
            "DEPARTMENT",
            f"NIGHT\n{self.night_date.strftime('%d/%m/%y') if self.night_date else ''}",
            f"DAY\n{self.day_date.strftime('%d/%m/%y') if self.day_date else ''}",
            "TOTAL",
        ]

        row = 2
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = bold
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = sub_header_fill
            cell.border = thin

        # ================= DATA =================
        row += 1
        sr_no = 1
        for line in self.line_ids.sorted("sequence"):
            ws.cell(row=row, column=1, value=sr_no).alignment = center

            # ======== 🔴 CHANGE START: department + description ========
            dept_text = line.department_id.name or ""
            if line.description:
                dept_text = f"{dept_text}\n({line.description.strip()})"

            dept_cell = ws.cell(row=row, column=2, value=dept_text)
            dept_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

            # ======== 🔴 CHANGE START: accurate auto row height ========
            import math

            base_height = 22
            chars_per_line = 35  # based on column B width (50)

            total_chars = len(dept_text.replace("\n", ""))
            manual_lines = dept_text.count("\n") + 1
            wrapped_lines = math.ceil(total_chars / chars_per_line)

            required_lines = max(manual_lines, wrapped_lines)

            ws.row_dimensions[row].height = base_height * required_lines
            # ======== 🔴 CHANGE END ========

            # ======== 🔴 CHANGE END ========

            ws.cell(row=row, column=3, value=line.night_count).alignment = center
            ws.cell(row=row, column=4, value=line.day_count).alignment = center
            ws.cell(row=row, column=5, value=line.total_count).alignment = center

            for col in range(1, 6):
                ws.cell(row=row, column=col).border = thin

            sr_no += 1
            row += 1

        # ================= TOTAL ROW =================
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row=row, column=1, value="TOTAL").font = bold
        ws.cell(row=row, column=1).alignment = center
        ws.cell(row=row, column=1).fill = sub_header_fill

        ws.cell(row=row, column=3, value=self.total_night).font = bold
        ws.cell(row=row, column=4, value=self.total_day).font = bold
        ws.cell(row=row, column=5, value=self.total_attendance).font = bold

        for col in range(1, 6):
            ws.cell(row=row, column=col).border = thin
            ws.cell(row=row, column=col).alignment = center

        # ======== 🔴 CHANGE START: vertical border between B & C ========
        for r in range(1, row + 1):
            cell_b = ws.cell(row=r, column=2)
            cell_b.border = Border(
                left=cell_b.border.left,
                right=thin_right,
                top=cell_b.border.top,
                bottom=cell_b.border.bottom,
            )
        # ======== 🔴 CHANGE END ========

        # ================= COLUMN WIDTH =================
        ws.column_dimensions["A"].width = 11.45
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 16

        # ================= DOWNLOAD =================
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        attachment = self.env["ir.attachment"].create({
            "name": "Department_Attendance.xlsx",
            "type": "binary",
            "datas": base64.b64encode(buffer.read()),
            "res_model": self._name,
            "res_id": self.id,
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        })

        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}?download=true",
            "target": "self",
        }
