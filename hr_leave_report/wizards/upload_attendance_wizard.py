import base64
import io
import logging
import re
from datetime import datetime
from odoo import models, fields, api, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    _logger.error('Cannot import openpyxl - please install it: pip install openpyxl')


class UploadAttendanceWizard(models.TransientModel):
    _name = 'hr.leave.upload.attendance.wizard'
    _description = 'Upload Attendance Wizard'

    file = fields.Binary(string="Attendance File", required=True)
    filename = fields.Char(string="Filename")

    def action_upload(self):
        """Process the uploaded Excel file and create leave records"""
        self.ensure_one()

        if not self.file:
            raise UserError(_("Please upload a file."))

        _logger.info("=" * 80)
        _logger.info("STARTING ATTENDANCE UPLOAD PROCESS")
        _logger.info("=" * 80)

        # Decode and load the Excel file
        try:
            file_data = base64.b64decode(self.file)
            workbook = openpyxl.load_workbook(filename=io.BytesIO(file_data), data_only=True)
            _logger.info(f"Excel file loaded successfully. Available sheets: {workbook.sheetnames}")
        except Exception as e:
            _logger.error(f"Failed to load Excel file: {str(e)}")
            raise UserError(_("Invalid file! Please upload a valid Excel file.\nError: %s") % str(e))

        # Use the first worksheet in the workbook
        sheet = workbook.active
        sheet_name = workbook.sheetnames[0] if workbook.sheetnames else 'Unknown'
        _logger.info(f"Using worksheet: '{sheet_name}'")

        # Parse report month and year
        report_month = None
        report_year = None

        _logger.info("\n--- Searching for Date Range in first 10 rows ---")
        for r in range(1, 11):
            row_vals = []
            for cell in sheet[r]:
                if cell.value:
                    row_vals.append(str(cell.value))

            full_text = " ".join(row_vals)
            _logger.info(f"Row {r}: {full_text[:100]}")

            if "From Date" in full_text or "from date" in full_text.lower():
                _logger.info(f"Found 'From Date' in row {r}: {full_text}")
                # Extract dates in format YYYY-MM-DD
                date_matches = re.findall(r'(\d{4}-\d{2}-\d{2})', full_text)
                if date_matches:
                    try:
                        first_date = datetime.strptime(date_matches[0], '%Y-%m-%d')
                        report_month = first_date.month
                        report_year = first_date.year
                        _logger.info(f"Extracted Report Month: {report_month}, Year: {report_year}")
                    except ValueError as e:
                        _logger.warning(f"Failed to parse date: {e}")
                break

        if not report_month or not report_year:
            # Fallback to current month/year
            now = datetime.now()
            report_month = now.month
            report_year = now.year
            _logger.warning(f"Could not find date range. Using current month: {report_month}/{report_year}")

        # Find header row and day columns
        header_row_idx = None
        emp_code_col = None
        day_columns = {}  # {day_number: column_index}

        _logger.info("\n" + "=" * 80)
        _logger.info("DETAILED SHEET ANALYSIS - FIRST 20 ROWS")
        _logger.info("=" * 80)

        # Print ALL content of first 20 rows for debugging
        for row_idx in range(1, 21):
            row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
            _logger.info(f"\nRow {row_idx}:")
            for col_idx, cell in enumerate(row[:20]):  # First 20 columns
                if cell is not None and str(cell).strip():
                    _logger.info(f"  Column {col_idx}: '{cell}' (type: {type(cell).__name__})")

        _logger.info("\n" + "=" * 80)
        _logger.info("SEARCHING FOR HEADER ROW")
        _logger.info("=" * 80)

        for row_idx in range(1, 31):  # Search first 30 rows
            row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]

            # Check if this row contains "Employee Code" or similar
            found_emp_code = False
            for col_idx, cell in enumerate(row):
                if cell:
                    cell_str = str(cell).strip().lower()
                    if any(keyword in cell_str for keyword in
                           ['employee code', 'emp code', 'code', 'employee id', 'emp id', 'emp no', 'employee no']):
                        header_row_idx = row_idx
                        emp_code_col = col_idx
                        found_emp_code = True
                        _logger.info(f"*** FOUND HEADER ROW at row {row_idx}, column {col_idx} ***")
                        _logger.info(f"Header cell value: '{cell}'")
                        break

            if found_emp_code:
                _logger.info(f"Full header row {row_idx}: {[str(c) for c in row if c][:15]}")

                # Find Employee Code column
                for col_idx, cell in enumerate(row):
                    if cell and ("Employee Code" in str(cell) or str(cell).strip() == "Code"):
                        emp_code_col = col_idx
                        _logger.info(f"Employee Code column index: {col_idx}")
                        break

                # Check rows ABOVE and BELOW for day numbers
                _logger.info(f"\n--- Checking surrounding rows for day numbers ---")

                # Check 3 rows above
                for offset in range(-3, 0):
                    day_row_idx = row_idx + offset
                    if day_row_idx < 1:
                        continue
                    day_row = list(sheet.iter_rows(min_row=day_row_idx, max_row=day_row_idx, values_only=True))[0]
                    _logger.info(f"Row {day_row_idx} (offset {offset}): {[str(c) for c in day_row[:15] if c]}")

                    for col_idx_day, cell_value in enumerate(day_row):
                        if cell_value:
                            if isinstance(cell_value, int) and 1 <= cell_value <= 31:
                                day_columns[cell_value] = col_idx_day
                            elif isinstance(cell_value, str) and cell_value.strip().isdigit():
                                day_num = int(cell_value.strip())
                                if 1 <= day_num <= 31:
                                    day_columns[day_num] = col_idx_day

                if day_columns:
                    _logger.info(f"Found {len(day_columns)} day columns from rows above header")
                    _logger.info(f"Day columns: {sorted(day_columns.items())[:10]}")

                # If no day columns found, check the row 1-2 rows above
                if not day_columns:
                    day_row_idx = row_idx - 1
                    day_row = list(sheet.iter_rows(min_row=day_row_idx, max_row=day_row_idx, values_only=True))[0]

                    _logger.info(f"\n--- Parsing Day Numbers from row {day_row_idx} ---")
                    _logger.info(f"Full row content: {day_row[:20]}")

                    for col_idx_day, cell_value in enumerate(day_row):
                        if cell_value:
                            _logger.info(
                                f"  Column {col_idx_day}: value='{cell_value}', type={type(cell_value).__name__}")
                            # Check if it's a number (day of month)
                            if isinstance(cell_value, int) and 1 <= cell_value <= 31:
                                day_columns[cell_value] = col_idx_day
                                _logger.info(f"    -> Mapped as Day {cell_value}")
                            elif isinstance(cell_value, str) and cell_value.strip().isdigit():
                                day_num = int(cell_value.strip())
                                if 1 <= day_num <= 31:
                                    day_columns[day_num] = col_idx_day
                                    _logger.info(f"    -> Mapped as Day {day_num}")

                _logger.info(f"Total days found: {len(day_columns)}")
                break

        if not header_row_idx or emp_code_col is None:
            _logger.error("Could not find 'Employee Code' column in header")
            raise UserError(_("Could not find 'Employee Code' header. Please check the file format."))

        if not day_columns:
            _logger.error("Could not find day number columns")
            raise UserError(_("Could not find day columns in the attendance report."))

        # Initialize models
        employee_obj = self.env['hr.employee']
        leave_obj = self.env['hr.leave']
        leave_type_obj = self.env['hr.leave.type']

        # Leave type mapping
        type_mapping = {
            'CL': 'Casual Leave',
            'EL': 'Earned Leave',
            'SL': 'Sick Leave',
            'ML': 'Management Leave',
            'OD': 'On Duty',
            'UL': 'Unpaid',
            'L': 'Legal Leaves 2024',
        }

        created_count = 0
        skipped_count = 0
        error_count = 0

        _logger.info("\n" + "=" * 80)
        _logger.info("PROCESSING EMPLOYEE RECORDS")
        _logger.info("=" * 80)

        # Process data rows
        for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
            row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]

            # Get employee code
            emp_code_raw = row[emp_code_col] if emp_code_col < len(row) else None
            if not emp_code_raw:
                continue

            # Normalize employee code
            if isinstance(emp_code_raw, (int, float)):
                emp_code = str(int(emp_code_raw))
            else:
                emp_code = str(emp_code_raw).strip()

            # Skip empty codes, 'None', or legend rows (containing '=')
            if not emp_code or emp_code == 'None' or '=' in emp_code:
                continue

            _logger.info(f"\n--- Row {row_idx}: Employee Code '{emp_code}' ---")

            # Find employee
            employee = employee_obj.search([('employee_code', '=', emp_code)], limit=1)
            if not employee:
                employee = employee_obj.search([('barcode', '=', emp_code)], limit=1)

            if not employee:
                _logger.warning(f"Employee '{emp_code}' not found in system")
                error_count += 1
                continue

            _logger.info(f"Found employee: {employee.name} (ID: {employee.id})")

            # Collect all leave days for this employee first
            leave_days = []  # List of (date, status, leave_type_id, is_half_day)
            
            if not day_columns:
                _logger.warning(f"No day columns available to process for employee {employee.name}")
            else:
                _logger.info(f"Processing {len(day_columns)} day columns for {employee.name}")
            
            for day_num, col_idx in sorted(day_columns.items()):
                if col_idx >= len(row):
                    continue

                status = row[col_idx]
                if not status:
                    continue

                status = str(status).strip().upper()
                
                _logger.info(f"  Day {day_num}: Raw status = '{status}' (column {col_idx})")

                # Skip attendance statuses
                if status in ['P', 'W', 'H', 'A', 'WO', 'PRESENT', 'WEEKOFF', 'HOLIDAY', 'ABSENT']:
                    _logger.info(f"    Skipping attendance status: {status}")
                    continue

                # Parse date
                try:
                    current_date = datetime(report_year, report_month, day_num).date()
                except ValueError as e:
                    _logger.warning(f"Invalid date: {day_num}/{report_month}/{report_year} - {e}")
                    continue

                # Handle Half Day
                is_half_day = False
                leave_type_name = None

                if status == 'HD':
                    is_half_day = True
                    hd_type = leave_type_obj.search([('name', 'ilike', 'Half Day')], limit=1)
                    if hd_type:
                        leave_type_name = hd_type.name
                    else:
                        leave_type_name = 'Casual Leave'  # Fallback
                    _logger.info(f"    Half Day detected, using type: {leave_type_name}")
                else:
                    leave_type_name = type_mapping.get(status)

                if not leave_type_name:
                    _logger.warning(f"    Unknown status '{status}' - skipping")
                    continue

                # Find leave type
                if 'Unpaid' in leave_type_name:
                    leave_type = leave_type_obj.search([
                        '|', ('name', 'ilike', 'Unpaid'), ('name', 'ilike', 'Unpaid Leave')
                    ], limit=1)
                else:
                    leave_type = leave_type_obj.search([('name', 'ilike', leave_type_name)], limit=1)

                if not leave_type:
                    _logger.warning(f"    Leave Type '{leave_type_name}' not found in system")
                    error_count += 1
                    continue

                _logger.info(f"    Found leave: {status} on {current_date} -> {leave_type.name}")
                leave_days.append((current_date, status, leave_type.id, is_half_day))

            # Group consecutive leaves of the same type
            leaves_for_employee = 0  # Initialize to avoid UnboundLocalError
            if leave_days:
                _logger.info(f"\n--- Consolidating {len(leave_days)} leave days for {employee.name} ---")
                leave_days.sort(key=lambda x: x[0])  # Sort by date
                
                consolidated_leaves = []
                current_group = None
                
                for leave_date, status, leave_type_id, is_half_day in leave_days:
                    if current_group is None:
                        # Start new group
                        current_group = {
                            'start_date': leave_date,
                            'end_date': leave_date,
                            'leave_type_id': leave_type_id,
                            'status': status,
                            'is_half_day': is_half_day,
                            'days': 1
                        }
                    elif (leave_type_id == current_group['leave_type_id'] and 
                          is_half_day == current_group['is_half_day'] and
                          (leave_date - current_group['end_date']).days == 1):
                        # Extend current group (consecutive day with same type)
                        current_group['end_date'] = leave_date
                        current_group['days'] += 1
                    else:
                        # Save current group and start new one
                        consolidated_leaves.append(current_group)
                        current_group = {
                            'start_date': leave_date,
                            'end_date': leave_date,
                            'leave_type_id': leave_type_id,
                            'status': status,
                            'is_half_day': is_half_day,
                            'days': 1
                        }
                
                # Don't forget the last group
                if current_group:
                    consolidated_leaves.append(current_group)
                
                _logger.info(f"Consolidated into {len(consolidated_leaves)} leave record(s)")
                
                # Create consolidated leave records
                leaves_for_employee = 0
                for leave_group in consolidated_leaves:
                    start_date = leave_group['start_date']
                    end_date = leave_group['end_date']
                    leave_type_id = leave_group['leave_type_id']
                    status = leave_group['status']
                    is_half_day = leave_group['is_half_day']
                    num_days = leave_group['days']
                    
                    # Check for duplicates
                    existing = leave_obj.search([
                        ('employee_id', '=', employee.id),
                        ('request_date_from', '=', start_date),
                        ('request_date_to', '=', end_date),
                        ('holiday_status_id', '=', leave_type_id),
                        ('state', 'not in', ['refuse', 'cancel']),
                    ])

                    if existing:
                        _logger.info(f"  Duplicate found - skipping {status} from {start_date} to {end_date}")
                        skipped_count += 1
                        continue

                    # Create leave record
                    if start_date == end_date:
                        name = f'Bulk Import: {status} on {start_date}'
                    else:
                        name = f'Bulk Import: {status} from {start_date} to {end_date}'
                    
                    vals = {
                        'employee_id': employee.id,
                        'holiday_status_id': leave_type_id,
                        'request_date_from': start_date,
                        'request_date_to': end_date,
                        'name': name,
                        'state': 'confirm',
                    }

                    if is_half_day:
                        vals.update({
                            'request_unit_half': True,
                            'request_date_from_period': 'am',
                            'number_of_days': 0.5 * num_days,
                        })
                    else:
                        vals['number_of_days'] = float(num_days)

                    try:
                        new_leave = leave_obj.create(vals)
                        new_leave.action_approve()  # Auto-approve
                        if new_leave.state != 'validate':
                            new_leave.action_validate()  # Final validation if needed
                        
                        created_count += 1
                        leaves_for_employee += 1
                        _logger.info(f"  ✓ Created: {status} from {start_date} to {end_date} ({num_days} days) - ID: {new_leave.id}")
                    except Exception as e:
                        _logger.error(f"  ✗ Failed to create leave: {str(e)}")
                        error_count += 1

            if leaves_for_employee > 0:
                _logger.info(f"Total leaves created for {employee.name}: {leaves_for_employee}")

        # Final summary
        _logger.info("\n" + "=" * 80)
        _logger.info("UPLOAD SUMMARY")
        _logger.info("=" * 80)
        _logger.info(f"Successfully created: {created_count} leaves")
        _logger.info(f"Skipped (duplicates): {skipped_count}")
        _logger.info(f"Errors: {error_count}")
        _logger.info("=" * 80)

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _("Upload Complete"),
                'message': _(
                    "Created: %s leaves\n"
                    "Skipped: %s duplicates\n"
                    # "Errors: %s\n\n"
                    # "Check server logs for details."
                ) % (created_count, skipped_count),
                'type': 'success' if error_count == 0 else 'warning',
                'sticky': True,
            }
        }