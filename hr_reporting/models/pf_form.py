from odoo import models
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage,ImageOps
import base64

class HrCustomFormPf(models.Model):
    _inherit = "hr.custom.form.pf"

    def action_generate_excel_report(self):
        """Generate Excel report for Daily Permit Work"""
        output = BytesIO()
        wb = Workbook()
        ws = wb.active

        # Add company logo if it exist
        if self.env.user.company_id.logo:
            max_width = 150
            max_height = 100
            image_data = base64.b64decode(self.env.user.company_id.logo)
            image = PILImage.open(BytesIO(image_data))
            width, height = image.size
            aspect_ratio = width / height

            if width > max_width:
                width = max_width
                height = int(width / aspect_ratio)

            if height > max_height:
                height = max_height
                width = int(height * aspect_ratio)

            padding_top = 10
            padding_left = 10

            resized_image = image.resize((width, height), PILImage.LANCZOS)
            ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='rgba(0,0,0,0)')
            img_bytes = BytesIO()
            resized_image.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            logo_image = Image(img_bytes)
            ws.add_image(logo_image, 'A1') # save image in cell 'A1'

        # Define Formating
        border = Border(top=Side(style='thin'), left=Side(style='thin'),right=Side(style='thin'), bottom=Side(style='thin'))
        white_side = Side(border_style="thin", color="FFFFFF")  # white color
        white_border = Border(left=white_side, right=white_side, top=white_side, bottom=white_side)
        align_center = Alignment(horizontal='center', vertical='center',wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='center', wrapText=True)
        font_header = Font(name='Arial', size=20, bold=True)
        font_title = Font(name='Arial', size=11, bold=True)
        font_all = Font(name='Arial', size=11, bold=False)
        white_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid') 
        teal_fill = PatternFill(start_color='005A70', end_color='005A70', fill_type='solid') 
        red_font = Font(name='Arial',size=11,bold=True,color="FF0000")


        # Column widths
        for i in range(1,27):
            col = get_column_letter(i)
            ws.column_dimensions[col].width = 15


        # ======= ROW 1 : Company Name =======
        ws.merge_cells('A1:Z1')
        ws['A1'] = "NEW EMPLOYEE REQUIRED DETAILS FOR UAN GENERATION"
        ws['A1'].font = font_header
        ws['A1'].alignment = align_center
        ws.row_dimensions[1].height = 80  # Logo
       

        # =========================== Hepler Function ===================================
        def write_cell(ws, row, col_start, col_end, value=None, fill=None, font=None, align=None, height=27,border=None):
            # merge cell
            ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
            cell = ws.cell(row=row, column=col_start)

            if value is not None:
                cell.value = value
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if align:
                cell.alignment = align
            if height:
                ws.row_dimensions[row].height = height
            if border:
                for col in range(col_start, col_end + 1):
                    ws.cell(row=row, column=col).border = border

        # ============================= End Function ====================================
        

        # ============================= Data Filling ====================================
        cur_row = 2

        # ============================= PF Details ====================================
        # Add heading text 
        write_cell(ws, cur_row, 1, 26,"PF Details",font=white_font,fill=teal_fill)
    
        cur_row += 1
        # sub-heading 
        write_cell(ws, cur_row, 1, 1,"Sr. No",align=align_center,font=font_title,fill=grey_fill)
        write_cell(ws, cur_row, 2, 3,"Old UAN No. (if  new Employees Have)",align=align_center,font=red_font,fill=grey_fill)
        write_cell(ws, cur_row, 4, 5,"DOL of Previous Employment",align=align_center,font=font_title,fill=grey_fill)
        write_cell(ws, cur_row, 6, 6,"Personal Title (Mr. /Ms, Mrs)",align=align_center,font=red_font,fill=grey_fill)
        write_cell(ws, cur_row, 7, 8,"Name",align=align_center,font=red_font,fill=grey_fill)
        write_cell(ws, cur_row, 9, 9,"Gender",align=align_center,font=red_font,fill=grey_fill)
        write_cell(ws, cur_row, 10, 10,"Date of Birth",align=align_center,font=red_font,fill=grey_fill)
        write_cell(ws, cur_row, 11, 11,"Father's/Husband Name",align=align_center,font=red_font,fill=grey_fill)
        write_cell(ws, cur_row, 12, 12,"Relation",align=align_center,font=red_font,fill=grey_fill)
        write_cell(ws, cur_row, 13, 13,"Marital Status",align=align_center,font=red_font,fill=grey_fill)
        write_cell(ws, cur_row, 14, 14,"Date of Joining",align=align_center,font=red_font,fill=grey_fill)
        write_cell(ws, cur_row, 15, 15,"Mobile Number",align=align_center,font=font_title,fill=grey_fill)
        write_cell(ws, cur_row, 16, 16,"PAN NUMBER",align=align_center,font=font_title,fill=grey_fill)
        write_cell(ws, cur_row, 17, 18,"AS PER PAN CARD NAME",align=align_center,font=font_title,fill=grey_fill)
        write_cell(ws, cur_row, 19, 20,"AADHAR NUMBER",align=align_center,font=red_font,fill=grey_fill)
        write_cell(ws, cur_row, 21, 22,"NAME AS PER AADHAR",align=align_center,font=red_font,fill=grey_fill)
        write_cell(ws, cur_row, 23, 24,"PF WAGES /PF SALARY /PF DECDUCTIBLE SALARY",align=align_center,font=red_font,fill=grey_fill)
        write_cell(ws, cur_row, 25, 26,"Is Pension Contribution Applicable Yes Or No",align=align_center,font=red_font,fill=grey_fill,height=45)
        
        cur_row += 1
        
        for rec in self.pf_line_ids:
            write_cell(ws, cur_row, 1, 1,rec.sequence or "")
            write_cell(ws, cur_row, 2, 3,rec.old_uan_no or "")
            write_cell(ws, cur_row, 4, 5,rec.dol_prev_employment or "")
            write_cell(ws, cur_row, 6, 6,rec.personal_title or "")
            write_cell(ws, cur_row, 7, 8,rec.employee_id.name or "")
            write_cell(ws, cur_row, 9, 9,rec.gender or "")
            write_cell(ws, cur_row, 10, 10,rec.date_of_birth.strftime('%d-%m-%Y') or '')
            write_cell(ws, cur_row, 11, 11,rec.father_husband_name or "")
            write_cell(ws, cur_row, 12, 12,rec.relation or "")
            write_cell(ws, cur_row, 13, 13,rec.marital_status or "")
            write_cell(ws, cur_row, 14, 14,rec.date_of_joining.strftime('%d-%m-%Y') or '')
            write_cell(ws, cur_row, 15, 15,rec.mobile_number or "")
            write_cell(ws, cur_row, 16, 16,rec.pan_number or "")
            write_cell(ws, cur_row, 17, 18,rec.pan_card_name or "")
            write_cell(ws, cur_row, 19, 20,rec.aadhaar_number or "")
            write_cell(ws, cur_row, 21, 22,rec.aadhaar_name or "")
            write_cell(ws, cur_row, 23, 24,rec.pf_wage_type or "")
            write_cell(ws, cur_row, 25, 26,rec.pension_applicable or "")
            
            cur_row += 1
            
        # Add one blank line
        write_cell(ws, cur_row, 1, 26)
        
        # ============================= End Data Filling =============================
        
        # add formating in all rows
        for row in range(1, cur_row + 1):
            for col_idx in range(1, 27):
                col = get_column_letter(col_idx)
                cell = ws[f"{col}{row}"]

                if cell.font == Font():
                    cell.font = font_all
                if cell.border == white_border:
                    continue
                else:
                    cell.border = border
                if cell.alignment == Alignment():
                    cell.alignment = align_left
                    
        # ======================== Notes ========================
        # add blank row
        cur_row += 1
        write_cell(ws, cur_row, 1, 10,"",border=white_border)
        
        cur_row += 1
        write_cell(ws, cur_row, 1, 1,"",border=white_border)
        write_cell(ws, cur_row, 2, 10,"Note",font=font_header,border=white_border)
        
        cur_row += 1
        write_cell(ws, cur_row, 2, 10,"*In Red Fonts is Mandatory",border=white_border,font=red_font)
        
        cur_row += 1
        write_cell(ws, cur_row, 1, 1,"1.",border=white_border,font=font_title,align=align_center)
        write_cell(ws, cur_row, 2, 10,"Please send the Aadhar card of the employees too.",font=red_font,border=white_border)
        
        cur_row += 1
        write_cell(ws, cur_row, 1, 1,"2.",border=white_border,font=font_title,align=align_center)
        write_cell(ws, cur_row, 2, 10,"If employees has Old UAN in that case we need member PF passbook for Previous establishment.",font=red_font,border=white_border)
        
        cur_row += 1
        write_cell(ws, cur_row, 1, 1,"3.",border=white_border,font=font_title,align=align_center)
        write_cell(ws, cur_row, 2, 10,"The date of joining, Date of Birth needs to mention in 11.04.2018 format. Don't use slash.",font=font_title,border=white_border)
        
        cur_row += 1
        write_cell(ws, cur_row, 1, 1,"4.",border=white_border,font=font_title,align=align_center)
        write_cell(ws, cur_row, 2, 10,"Please send the above data within 7 days of joining of employee.",font=font_title,border=white_border)
        
        cur_row += 1
        write_cell(ws, cur_row, 1, 10,"",border=white_border)
                    
        # Save workbook
        wb.save(output)
        output.seek(0)

        xls_file = base64.b64encode(output.read())
        output.close()

        # Create attachment
        filename = (f"PF Form ({self.name or ''}).xls")
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': xls_file,
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        # Return download action
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }