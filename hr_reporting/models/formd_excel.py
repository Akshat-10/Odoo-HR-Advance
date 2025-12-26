from odoo import models, fields
from io import BytesIO
import base64
import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from PIL import Image as PILImage
from openpyxl.drawing.image import Image
from odoo.tools.misc import format_date



class HrCustomFormFormD(models.Model):
    _inherit = "hr.custom.form.formd"

    generate_xls_file = fields.Binary(string="Form D Excel")

    def action_generate_formd_excel(self):
        self.ensure_one()

        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Form D Register"

        # LOGO
        if self.env.user.company_id.logo:
            max_width, max_height = 200, 80
            image_data = base64.b64decode(self.env.user.company_id.logo)
            image = PILImage.open(io.BytesIO(image_data))
            width, height = image.size
            aspect_ratio = width / height

            if width > max_width:
                width = max_width
                height = int(width / aspect_ratio)
            if height > max_height:
                height = max_height
                width = int(height * aspect_ratio)

            resized_image = image.resize((width, height), PILImage.LANCZOS)
            img_bytes = io.BytesIO()
            resized_image.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            ws.add_image(Image(img_bytes), 'A1')

        # ================= STYLES =================
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

        title_font = Font(size=14, bold=True)
        header_font = Font(size=12,bold=True)

        blue_fill = PatternFill(start_color='87CEFA', end_color='87CEFA', fill_type='solid')
        sub_blue_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
        grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

        ws.row_dimensions[1].height = 60
        ws.row_dimensions[2].height = 40
        ws.row_dimensions[3].height = 15
        ws.row_dimensions[4].height = 15

        # ================= ROW 1 =================
        ws.merge_cells("A1:J1")
        ws["A1"] = 'GERMAN TMT PRIVATE LIMITED'
        ws["A1"].font = Font(size=20, bold=True)
        ws["A1"].fill = blue_fill
        ws["A1"].alignment = align_center

        # ================= ROW 2 =================
        ws.merge_cells("A2:J2")
        ws["A2"] = "FORM - D"
        ws["A2"].font = Font(size=18, bold=True)
        ws["A2"].fill = blue_fill
        ws["A2"].alignment = align_center

        # ================= ROW 3 =================
        ws["A3"] = "Company"
        ws["A3"].font = header_font
        ws["A3"].alignment = align_left
        ws["A3"].fill = grey_fill

        ws.merge_cells("B3:C3")
        ws["B3"] = self.company_id.name or ""
        ws["B3"].alignment = align_left

        ws["D3"] = "Date"
        ws["D3"].font = header_font
        ws["D3"].alignment = align_left
        ws["D3"].fill = grey_fill

        ws.merge_cells("E3:F3")
        formatted_date = format_date(self.env, fields.Date.context_today(self))
        ws["E3"] = formatted_date
        ws["E3"].alignment = align_left

        ws.merge_cells("G3:H3")
        ws["G3"] = "Month and Year"
        ws["G3"].font = header_font
        ws["G3"].alignment = align_left
        ws["G3"].fill = grey_fill

        ws.merge_cells("I3:J3")
        ws["I3"] = self.month_year or ""
        ws["I3"].alignment = align_left

        # ================= ROW 4 =================
        ws["A4"] = "Total Number of Workers"
        ws["A4"].font = header_font
        ws["A4"].alignment = align_left
        ws["A4"].fill = grey_fill

        ws.merge_cells("B4:C4")
        ws["B4"] = self.total_workers or 0
        ws["B4"].alignment = align_left

        # Document Reference
        ws.merge_cells("D4:E4")
        ws["D4"] = "Document Reference"
        ws["D4"].font = header_font
        ws["D4"].alignment = align_left
        ws["D4"].fill = grey_fill

        ws.merge_cells("F4:J4")
        ws["F4"] = self.name or ""
        ws["F4"].alignment = align_left

        for row in [1, 2, 3, 4]:
            for col in range(1, 11):
                ws.cell(row=row, column=col).border = border
                # ws.cell(row=row, column=col).alignment = align_left


        ws.merge_cells("A5:A6")
        ws.merge_cells("B5:B6")
        ws.merge_cells("C5:C6")
        ws.merge_cells("D5:D6")
        ws.merge_cells("E5:E6")
        ws.merge_cells("F5:F6")
        ws.merge_cells("G5:J5")

        ws["A5"] = "Category of Workers"
        ws["B5"] = "Brief Description of Work"
        ws["C5"] = "No. of men employed"
        ws["D5"] = "No. of women employed"
        ws["E5"] = "Rate of remuneration paid"
        ws["F5"] = "Basic Wages or Salary"
        ws["G5"] = "Parts of Wages"

        ws["G6"] = "D.A"
        ws["H6"] = "H.R.A"
        ws["I6"] = "Other Allowances"
        ws["J6"] = "Cash Value of concessional\nSupply"

        for row in [5, 6]:
            for col in range(1, 11):
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.alignment = align_center
                cell.fill = sub_blue_fill
                cell.border = border

        # ================= DATA ROWS =================
        row = 7
        for line in self.line_ids:
            ws.cell(row=row, column=1, value=line.category_id.name if line.category_id else "")
            ws.cell(row=row, column=2, value=line.work_description or "")
            ws.cell(row=row, column=3, value=line.men_employed or 0)
            ws.cell(row=row, column=4, value=line.women_employed or 0)
            ws.cell(row=row, column=5, value=line.remuneration_rate or "")
            ws.cell(row=row, column=6, value=line.basic_wages or 0.0)
            ws.cell(row=row, column=7, value=line.part_da or 0.0)
            ws.cell(row=row, column=8, value=line.part_hra or 0.0)
            ws.cell(row=row, column=9, value=line.part_other_allowances or 0.0)
            ws.cell(row=row, column=10, value=line.part_cash_value or 0.0)

            for col in range(1, 11):
                ws.cell(row=row, column=col).border = border
                ws.cell(row=row, column=col).alignment = align_center

            row += 1

        # ================= COLUMN WIDTH =================
        widths = [25, 32, 12, 14, 22, 20, 12, 12, 20, 28]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w

        wb.save(output)
        output.seek(0)

        self.generate_xls_file = base64.b64encode(output.getvalue())

        document_reference = (
            self.name.strip()
            .replace(" ", "_")
            .replace("/", "_")
            if self.name
            else "Document_Reference"
        )

        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/?model=hr.custom.form.formd&id={self.id}"
                   f"&field=generate_xls_file"
                   f"&filename=Form_D_Register_{document_reference}.xlsx"
                   f"&download=true",
            "target": "self",
        }

