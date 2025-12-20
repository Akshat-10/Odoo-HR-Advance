# -*- coding: utf-8 -*-
from odoo import models, fields
from io import BytesIO
import base64
import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from odoo.exceptions import UserError


class HrSalaryAttachment(models.Model):
    _inherit = 'hr.salary.attachment'

    salary_excel_file = fields.Binary(string="Salary Sheet Excel")
    salary_excel_filename = fields.Char(string="Excel Filename")

    def action_generate_salary_attachment_excel(self):
        records = self.env['hr.salary.attachment'].browse(
            self.env.context.get('active_ids', self.ids)
        )

        if not records:
            raise UserError("No records selected.")

        def apply_border(ws, cell_range, border):
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = border

        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Salary Sheet"

        # =================================================
        # COLLECT ALL ATTACHMENT TYPES
        # =================================================
        attachment_types = []

        for rec in records:
            lines = None
            if hasattr(rec, 'attachment_line_ids'):
                lines = rec.attachment_line_ids
            elif hasattr(rec, 'line_ids'):
                lines = rec.line_ids
            elif hasattr(rec, 'salary_attachment_line_ids'):
                lines = rec.salary_attachment_line_ids

            if lines:
                for line in lines:
                    type_name = None
                    if hasattr(line, 'salary_attachment_type_id') and line.salary_attachment_type_id:
                        type_name = line.salary_attachment_type_id.name
                    elif hasattr(line, 'type_id') and line.type_id:
                        type_name = line.type_id.name
                    elif hasattr(line, 'attachment_type_id') and line.attachment_type_id:
                        type_name = line.attachment_type_id.name

                    if type_name and type_name not in attachment_types:
                        attachment_types.append(type_name)

        rec0 = records[0]

        # =================================================
        # LOGO
        # =================================================
        if rec0.company_id.logo:
            max_width, max_height = 200, 80
            image_data = base64.b64decode(rec0.company_id.logo)
            image = PILImage.open(io.BytesIO(image_data))
            image.thumbnail((max_width, max_height))
            img_bytes = io.BytesIO()
            image.save(img_bytes, format="PNG")
            img_bytes.seek(0)
            ws.add_image(Image(img_bytes), "A1")

        # =================================================
        # STYLES
        # =================================================
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center")

        title_font = Font(size=20, bold=True)
        sub_title_font = Font(size=18, bold=True)
        header_font = Font(size=12, bold=True)

        # 🔴 ADDED (required for total row)
        total_font = Font(bold=True, color="FF0000")

        blue_fill = PatternFill(start_color="87CEFA", end_color="87CEFA", fill_type="solid")
        sub_blue_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        ws.row_dimensions[1].height = 60
        ws.row_dimensions[2].height = 40

        base_widths = [18, 30, 22, 25, 25, 18, 18]
        all_widths = base_widths + [15] * len(attachment_types)

        for i, w in enumerate(all_widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w

        total_columns = 7 + len(attachment_types)
        last_col_letter = chr(64 + total_columns)

        # =================================================
        # HEADER
        # =================================================
        ws.merge_cells(f"A1:{last_col_letter}1")
        ws["A1"] = rec0.company_id.name or ""
        ws["A1"].font = title_font
        ws["A1"].alignment = align_center
        ws["A1"].fill = blue_fill
        apply_border(ws, f"A1:{last_col_letter}1", border)

        ws.merge_cells(f"A2:{last_col_letter}2")
        ws["A2"] = "TOTAL ADVANCE"
        ws["A2"].font = sub_title_font
        ws["A2"].alignment = align_center
        ws["A2"].fill = blue_fill
        apply_border(ws, f"A2:{last_col_letter}2", border)

        ws["A3"] = "Date"
        ws["A3"].fill = grey_fill
        ws.merge_cells("B3:C3")
        ws["A3"].alignment = align_left
        ws["B3"] = fields.Date.context_today(self)

        ws["D3"] = "Month & Year"
        ws["D3"].fill = grey_fill
        ws.merge_cells(f"E3:{last_col_letter}3")
        date_val = getattr(rec0, "date_start", False) or getattr(rec0, "date_from", False)
        ws["E3"] = date_val.strftime("%B %Y") if date_val else ""
        apply_border(ws, f"A3:{last_col_letter}3", border)

        # =================================================
        # TABLE HEADER
        # =================================================
        base_headers = [
            "Employee Code",
            "Employee Name",
            "Department",
            "Designation",
            "Description",
            "Start Date",
            "Total Amount",
        ]
        all_headers = base_headers + attachment_types

        for col, text in enumerate(all_headers, start=1):
            cell = ws.cell(row=4, column=col, value=text)
            cell.font = header_font
            cell.fill = sub_blue_fill
            cell.alignment = align_center
            cell.border = border

        # =================================================
        # DATA ROWS
        # =================================================
        row = 5

        # 🔴 ADDED (required for total row)
        total_amount_sum = 0.0
        type_totals = {t: 0.0 for t in attachment_types}

        for rec in records:
            attachment_data = {}

            lines = None
            if hasattr(rec, 'attachment_line_ids'):
                lines = rec.attachment_line_ids
            elif hasattr(rec, 'line_ids'):
                lines = rec.line_ids
            elif hasattr(rec, 'salary_attachment_line_ids'):
                lines = rec.salary_attachment_line_ids

            if lines:
                for line in lines:
                    type_name = None
                    if hasattr(line, 'salary_attachment_type_id') and line.salary_attachment_type_id:
                        type_name = line.salary_attachment_type_id.name
                    elif hasattr(line, 'type_id') and line.type_id:
                        type_name = line.type_id.name
                    elif hasattr(line, 'attachment_type_id') and line.attachment_type_id:
                        type_name = line.attachment_type_id.name

                    amount = line.amount if hasattr(line, 'amount') else line.payslip_amount or 0.0
                    if type_name:
                        attachment_data[type_name] = amount
                        type_totals[type_name] += amount

            for employee in rec.employee_ids:
                employee_code = (
                    employee.registration_number or
                    employee.barcode or
                    employee.employee_code or
                    employee.identification_id or ""
                )

                ws.cell(row=row, column=1, value=employee_code)
                ws.cell(row=row, column=2, value=employee.name or "")
                ws.cell(row=row, column=3, value=employee.department_id.name if employee.department_id else "")
                ws.cell(row=row, column=4, value=employee.job_id.name if employee.job_id else "")
                ws.cell(row=row, column=5, value=rec.description or "")
                ws.cell(row=row, column=6, value=rec.date_start)
                ws.cell(row=row, column=7, value=rec.total_amount or 0.0)

                total_amount_sum += rec.total_amount or 0.0

                for idx, t in enumerate(attachment_types):
                    ws.cell(row=row, column=8 + idx, value=attachment_data.get(t, 0.0))

                for c in range(1, total_columns + 1):
                    ws.cell(row=row, column=c).border = border
                    ws.cell(row=row, column=c).alignment = align_center

                row += 1

        # =================================================
        # FINAL TOTAL ROW (LIKE YOUR SCREENSHOT)
        # =================================================
        row += 1
        ws.cell(row=row, column=6, value="SUB TOTAL").font = header_font
        ws.cell(row=row, column=6).alignment = align_center

        ws.cell(row=row, column=7, value=total_amount_sum).font = total_font

        for idx, t in enumerate(attachment_types):
            ws.cell(row=row, column=8 + idx, value=type_totals.get(t, 0.0)).font = total_font

        for c in range(1, total_columns + 1):
            ws.cell(row=row, column=c).border = border
            ws.cell(row=row, column=c).alignment = align_center

        wb.save(output)
        output.seek(0)

        rec0.salary_excel_file = base64.b64encode(output.read())
        rec0.salary_excel_filename = "Total_Advance.xlsx"

        return {
            "type": "ir.actions.act_url",
            "url": (
                "/web/content/?model=hr.salary.attachment"
                f"&id={rec0.id}"
                "&field=salary_excel_file"
                "&filename_field=salary_excel_filename"
                "&download=true"
            ),
            "target": "self",
        }
