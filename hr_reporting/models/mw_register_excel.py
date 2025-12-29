from odoo import models, fields
from io import BytesIO
import base64, io
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from odoo.tools.misc import format_date



class HrCustomFormMwNotice(models.Model):
    _inherit = "hr.custom.form.mw_notice"

    generate_xls_file = fields.Binary(string="MV Register Excel")

    def action_generate_mv_register_excel(self):
        self.ensure_one()

        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Minimum Wages Register"

        # LOGO
        if self.env.user.company_id.logo:
            max_width, max_height = 200, 80
            image_data = base64.b64decode(self.env.user.company_id.logo)
            image = PILImage.open(io.BytesIO(image_data))
            width, height = image.size
            aspect_ratio = width / height

            if width > max_width:
                width = max_width
                height = int(width / aspect_ratio)
            if height > max_height:
                height = max_height
                width = int(height * aspect_ratio)

            resized_image = image.resize((width, height), PILImage.LANCZOS)
            img_bytes = io.BytesIO()
            resized_image.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            ws.add_image(Image(img_bytes), 'A1')

        # STYLES
        border = Border(
            top=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin'),
            bottom=Side(style='thin')
        )
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        font_header = Font(name='Arial', size=12, bold=True)
        font_regular = Font(name='Arial', size=11)
        blue_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
        light_blue_fill = PatternFill(start_color='87CEFA', end_color='87CEFA', fill_type='solid')
        grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

        ws.row_dimensions[1].height = 60
        ws.row_dimensions[2].height = 40
        ws.row_dimensions[3].height = 25


        ws.merge_cells("A1:O1")
        ws["A1"] = 'GERMAN TMT PRIVATE LIMITED'
        ws["A1"].font = Font(size=20, bold=True)
        ws["A1"].fill = light_blue_fill
        ws["A1"].alignment = align_center


        ws.merge_cells("A2:O2")
        ws["A2"] = "Minimum Wages Register"
        ws["A2"].font = Font(size=18, bold=True)
        ws["A2"].fill = light_blue_fill
        ws["A2"].alignment = align_center

        # 🔹 BORDER FIX FOR ROW 1 & 2
        for col in range(1, 16):
            ws.cell(row=1, column=col).border = border
            ws.cell(row=2, column=col).border = border

        #  ROW 3 (DOCUMENT INFO)
        ws.merge_cells("A3:B3")
        ws.merge_cells("C3:E3")
        ws.merge_cells("F3:F3")
        ws.merge_cells("G3:I3")
        ws.merge_cells("J3:J3")
        ws.merge_cells("K3:O3")

        ws["A3"] = "Document Reference"
        ws["C3"] = self.name or ""

        ws["F3"] = "Company"
        ws["G3"] = self.company_id.name or ""

        ws["J3"] = "Company Address"
        ws["K3"] = self.company_id.partner_id._display_address() if self.company_id.partner_id else ""

        for cell in ["A3", "F3", "J3"]:
            ws[cell].font = font_header
            ws[cell].alignment = align_left
            ws[cell].fill = grey_fill

        for cell in ["C3", "G3", "K3"]:
            ws[cell].font = font_regular
            ws[cell].alignment = align_left

        for col in range(1, 16):
            ws.cell(row=3, column=col).border = border

        #  HEADERS
        headers = [
            "SR.NO", "EMPLOYEE", "FATHER / HUSBAND NAME", "GENDER",
            "DEPARTMENT", "ABSENCE DATE", "DAMAGE / LOSS",
            "DAMAGE DATE", "SHOWED CAUSE", "SHOW CAUSE DATE",
            "DEDUCTION DATE", "DEDUCTION AMOUNT",
            "INSTALLMENTS", "REALISATION DATE", "REMARKS"
        ]

        header_row = 4
        for col, title in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=title)
            cell.font = font_header
            cell.fill = blue_fill
            cell.alignment = align_center
            cell.border = border

        # data
        row = header_row + 1
        sr_no = 1

        for line in self.notice_line_ids:
            ws.cell(row=row, column=1, value=sr_no).alignment = align_center
            ws.cell(row=row, column=2, value=line.employee_id.name or '').alignment = align_left
            ws.cell(row=row, column=3, value=line.father_name or '').alignment = align_left
            ws.cell(
                row=row,
                column=4,
                value=dict(line._fields['gender'].selection).get(line.gender, '') if line.gender else ''
            ).alignment = align_center
            ws.cell(row=row, column=5, value=line.department_id.name or '').alignment = align_left
            ws.cell(row=row, column=6, value=format_date(self.env, line.absence_date) if line.absence_date else "").alignment = align_center
            ws.cell(row=row, column=7, value=line.damages_description or '').alignment = align_left
            ws.cell(row=row, column=8, value=format_date(self.env, line.damages_date) if line.damages_date else "").alignment = align_center
            ws.cell(
                row=row,
                column=9,
                value=dict(line._fields['showed_cause'].selection).get(line.showed_cause, '') if line.showed_cause else ''
            ).alignment = align_center
            ws.cell(row=row, column=10, value=format_date(self.env, line.show_cause_date) if line.show_cause_date else "").alignment = align_center
            ws.cell(row=row, column=11, value=format_date(self.env, line.deduction_date) if line.deduction_date else "").alignment = align_center
            ws.cell(row=row, column=12, value=line.deduction_amount or 0).alignment = align_center
            ws.cell(row=row, column=13, value=line.installment_count or 0).alignment = align_center
            ws.cell(row=row, column=14, value=format_date(self.env, line.realisation_date) if line.realisation_date else "").alignment = align_center
            ws.cell(row=row, column=15, value=line.remarks or '').alignment = align_left

            for col in range(1, 16):
                ws.cell(row=row, column=col).border = border
                ws.cell(row=row, column=col).font = font_regular

            sr_no += 1
            row += 1

        # dynamic row
        data_rows_count = len(self.notice_line_ids) if self.notice_line_ids else 0
        if data_rows_count < 20:
            empty_rows_needed = 20 - data_rows_count
        else:
            empty_rows_needed = 10

        for _ in range(empty_rows_needed):
            for col_num in range(1, 16):  # A to O
                ws.cell(row=row, column=col_num, value="").alignment = align_left
                ws.cell(row=row, column=col_num).border = border
            row += 1


        widths = [8, 18, 26, 13, 18, 19, 26, 18, 20, 24, 19, 19, 20, 19, 20]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = width

        wb.save(output)
        output.seek(0)



        self.generate_xls_file = base64.b64encode(output.getvalue())

        # ========= FILE NAME LOGIC =========
        document_reference = (
            self.name.strip().replace(" ", "_")
            if self.name
            else "Document_Reference"
        )

        return {
            "type": "ir.actions.act_url",
            "url": (
                f"/web/content/?model=hr.custom.form.mw_notice&id={self.id}"
                f"&field=generate_xls_file"
                f"&filename=MV_Register_{document_reference}.xlsx"
                f"&download=true"
            ),
            "target": "self",
        }




