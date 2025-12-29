from odoo import models
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage,ImageOps
import base64

class LeaveApplication(models.Model):
    _inherit = "hr.custom.form.esic_declaration"

    def action_generate_excel_report(self):
        """Generate Excel report for Daily Permit Work"""
        output = BytesIO()
        wb = Workbook()
        ws = wb.active

        # Add company logo if it exist
        if self.env.user.company_id.logo:
            max_width = 150
            max_height = 100
            image_data = base64.b64decode(self.env.user.company_id.logo)
            image = PILImage.open(BytesIO(image_data))
            width, height = image.size
            aspect_ratio = width / height

            if width > max_width:
                width = max_width
                height = int(width / aspect_ratio)

            if height > max_height:
                height = max_height
                width = int(height * aspect_ratio)

            padding_top = 10
            padding_left = 10

            resized_image = image.resize((width, height), PILImage.LANCZOS)
            ImageOps.expand(resized_image, border=(padding_left, padding_top, 0, 0), fill='rgba(0,0,0,0)')
            img_bytes = BytesIO()
            resized_image.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            logo_image = Image(img_bytes)
            ws.add_image(logo_image, 'A1') # save image in cell 'A1'

        # Define Formating
        border = Border(top=Side(style='thin'), left=Side(style='thin'),right=Side(style='thin'), bottom=Side(style='thin'))
        white_side = Side(border_style="thin", color="FFFFFF")  # white color
        white_border = Border(left=white_side, right=white_side, top=white_side, bottom=white_side)
        align_center = Alignment(horizontal='center', vertical='center',wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='center', wrapText=True)
        font_header = Font(name='Arial', size=20, bold=True)
        font_title = Font(name='Arial', size=11, bold=True)
        font_all = Font(name='Arial', size=11, bold=False)
        white_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid') 
        teal_fill = PatternFill(start_color='005A70', end_color='005A70', fill_type='solid') 
        red_font = Font(name='Arial',size=11,bold=True,color="FF0000")

        # Column widths
        for i in range(1,9):
            col = get_column_letter(i)
            ws.column_dimensions[col].width = 15


        # ======= ROW 1 : Company Name =======
        ws.merge_cells('A1:H1')
        ws['A1'] = "ESIC Declaration Form"
        ws['A1'].font = font_header
        ws['A1'].alignment = align_center
        ws.row_dimensions[1].height = 80  # Logo
       

        # =========================== Hepler Function ===================================
        def write_cell(ws, row, col_start, col_end, value=None, fill=None, font=None, align=None, height=27,border=None):
            # merge cell
            ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
            cell = ws.cell(row=row, column=col_start)

            if value is not None:
                cell.value = value
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if align:
                cell.alignment = align
            if height:
                ws.row_dimensions[row].height = height
            if border:
                for col in range(col_start, col_end + 1):
                    ws.cell(row=row, column=col).border = border

        # ============================= End Function ====================================

        # ============================= Data Filling ====================================
        cur_row = 2

        # Add heading text 
        write_cell(ws, cur_row, 1, 8,"Basic Details",font=white_font,fill=teal_fill,align=align_center)

        cur_row += 1
        # Add text and data for 'Employee Name'
        write_cell(ws, cur_row, 1, 2,"Employee Name",font=red_font,fill=grey_fill)
        write_cell(ws, cur_row, 3, 4,self.employee_id.name or "")
        # Add text and data of 'Date Of Birth'
        write_cell(ws, cur_row, 5, 6,"Date Of Birth",font=red_font,fill=grey_fill)
        write_cell(ws, cur_row, 7, 8,self.esic_date_of_birth.strftime('%d-%m-%Y') or '')
        
        cur_row += 1
        # Add text and data for 'Father name'
        write_cell(ws, cur_row, 1, 2,"Father name",font=red_font,fill=grey_fill)
        write_cell(ws, cur_row, 3, 4,self.esic_father_name or "")
        # Add text and data of 'Gender'
        write_cell(ws, cur_row, 5, 6,"Gender",font=red_font,fill=grey_fill)
        write_cell(ws, cur_row, 7, 8,self.esic_gender or "")
        
        cur_row += 1
        # Add text and data for 'Marital Status'
        write_cell(ws, cur_row, 1, 2,"Marital Status",font=red_font,fill=grey_fill)
        write_cell(ws, cur_row, 3, 8,self.esic_marital_status or "")
        
        cur_row += 1
        # Add text and data for 'Present Address'
        write_cell(ws, cur_row, 1, 2,"Present Address",font=red_font,fill=grey_fill)
        write_cell(ws, cur_row, 3, 4,self.esic_present_address or "")
        # Add text and data of 'Permanent Address'
        write_cell(ws, cur_row, 5, 6,"Permanent Address",font=font_title,fill=grey_fill)
        write_cell(ws, cur_row, 7, 8,self.esic_permanent_address or "")
        
        cur_row += 1
        # Add text and data for 'Date of Joining'
        write_cell(ws, cur_row, 1, 2,"Date of Joining",font=red_font,fill=grey_fill)
        write_cell(ws, cur_row, 3, 4,self.esic_date_of_joining.strftime('%d-%m-%Y') or '')
        # Add text and data of 'Mobile Number'
        write_cell(ws, cur_row, 5, 6,"Mobile Number",font=red_font,fill=grey_fill)
        write_cell(ws, cur_row, 7, 8,self.esic_mobile_number or "")
       
        
        # Add one blank line
        cur_row += 1
        write_cell(ws, cur_row, 1, 8)
        
        
        # ============================= Family Details ====================================
        cur_row += 1

        # Add heading text 
        write_cell(ws, cur_row, 1, 8,"Family Details",font=white_font,fill=teal_fill,align=align_center)
        

        # ---------- Header Row 1 ----------
        cur_row += 1
        
        # heading 'Sr. No'
        ws.merge_cells(start_row=cur_row,end_row=cur_row+1,start_column=1,end_column=1)
        ws.cell(row=cur_row, column=1).value = "Sr. No"
        ws.cell(row=cur_row, column=1).font = font_title
        ws.cell(row=cur_row, column=1).fill = grey_fill 
        
        
        # heading 'Family Member's name'
        ws.merge_cells(start_row=cur_row,end_row=cur_row+1,start_column=2,end_column=2)
        ws.cell(row=cur_row, column=2).value = "Family Member's name"
        ws.cell(row=cur_row, column=2).font = font_title
        ws.cell(row=cur_row, column=2).fill = grey_fill 
        
        # heading 'Relation With Employee'
        ws.merge_cells(start_row=cur_row,end_row=cur_row+1,start_column=3,end_column=3)
        ws.cell(row=cur_row, column=3).value = "Relation With Employee"
        ws.cell(row=cur_row, column=3).font = font_title
        ws.cell(row=cur_row, column=3).fill = grey_fill 
        
        # heading 'Date of Birth'
        ws.merge_cells(start_row=cur_row,end_row=cur_row+1,start_column=4,end_column=4)
        ws.cell(row=cur_row, column=4).value = "Date of Birth"
        ws.cell(row=cur_row, column=4).font = font_title
        ws.cell(row=cur_row, column=4).fill = grey_fill 
        
        # heading 'Whether Residing with IP?'
        ws.merge_cells(start_row=cur_row,end_row=cur_row+1,start_column=5,end_column=5)
        ws.cell(row=cur_row, column=5).value = "Whether Residing with IP?"
        ws.cell(row=cur_row, column=5).font = font_title
        ws.cell(row=cur_row, column=5).fill = grey_fill 
          
        # heading 'Aadhar No.'
        ws.merge_cells(start_row=cur_row,end_row=cur_row+1,start_column=8,end_column=8)
        ws.cell(row=cur_row, column=8).value = "Aadhar No."
        ws.cell(row=cur_row, column=8).font = font_title
        ws.cell(row=cur_row, column=8).fill = grey_fill 
        
        
        # heading 'If No, State Place of Residence'
        write_cell(ws, cur_row, 6, 7,"If No, State Place of Residence",font=font_title,fill=grey_fill,align=align_center)
        
        # sub-heading 'State & District'
        write_cell(ws, cur_row+1, 6, 6,"State",align=align_center,font=font_title,fill=grey_fill)
        write_cell(ws, cur_row+1, 7, 7,"District",align=align_center,font=font_title,fill=grey_fill)
        
        cur_row += 2
        
        for i, rec in enumerate(self.family_line_ids):
            write_cell(ws, cur_row, 1, 1,i+1 or "")
            write_cell(ws, cur_row, 2, 2,rec.family_member_name or "")
            write_cell(ws, cur_row, 3, 3,rec.relation or "")
            write_cell(ws, cur_row, 4, 4,rec.date_of_birth.strftime('%d-%m-%Y') or '')
            write_cell(ws, cur_row, 5, 5,rec.residing_with_ip or "")
            write_cell(ws, cur_row, 6, 6,rec.residence_state or "")
            write_cell(ws, cur_row, 7, 7,rec.residence_district or "")
            write_cell(ws, cur_row, 8, 8,rec.aadhaar_no or "")
            cur_row += 1
                
        # Add one blank line
        write_cell(ws, cur_row, 1, 8)
        
        # ============================= Nominee Details ====================================
        cur_row += 1
        # Add heading text 
        write_cell(ws, cur_row, 1, 8,"Nominee Details",font=white_font,fill=teal_fill,align=align_center)
    
        cur_row += 1
        # sub-heading 'Name of Nominee'
        write_cell(ws, cur_row, 1, 2,"Name of Nominee",align=align_center,font=red_font,fill=grey_fill)
        
        # sub-heading 'Relation With Employee'
        write_cell(ws, cur_row, 3, 4,"Relation With Employee",align=align_center,font=red_font,fill=grey_fill)
        
        # sub-heading 'Address of Nominee'
        write_cell(ws, cur_row, 5, 6,"Address of Nominee",align=align_center,font=red_font,fill=grey_fill)
        
        # sub-heading 'Contact No.'
        write_cell(ws, cur_row, 7, 8,"Contact No.",align=align_center,font=red_font,fill=grey_fill)
        
        
        cur_row += 1
        
        for rec in self.nominee_line_ids:
            write_cell(ws, cur_row, 1, 2,rec.nominee_name or "")
            write_cell(ws, cur_row, 3, 4,rec.relation or "")
            write_cell(ws, cur_row, 5, 6,rec.nominee_address or "")
            write_cell(ws, cur_row, 7, 8,rec.contact_no or "")
            cur_row += 1
            
        # Add one blank line
        write_cell(ws, cur_row, 1, 8)
        
        # ============================= Employee Bank Details Compulsory ====================================
        cur_row += 1

        # Add heading text 
        write_cell(ws, cur_row, 1, 8,"Employee Bank Details Compulsory",font=white_font,fill=teal_fill,align=align_center)
    
        cur_row += 1
        # Name of Bank:
        write_cell(ws, cur_row, 1, 3,"Name of Bank:",font=red_font,fill=grey_fill)
        write_cell(ws, cur_row, 4, 8,self.esic_bank_name or "") 
        
        cur_row += 1
        # Bank Account Number
        write_cell(ws, cur_row, 1, 3,"Bank Account Number",font=red_font,fill=grey_fill)
        write_cell(ws, cur_row, 4, 8,self.esic_bank_account_no or "") 
        
        cur_row += 1
        # IFSC Code
        write_cell(ws, cur_row, 1, 3,"IFSC Code",font=red_font,fill=grey_fill)
        write_cell(ws, cur_row, 4, 8,self.esic_bank_ifsc or "") 
        
        cur_row += 1
        # MICR Number
        write_cell(ws, cur_row, 1, 3,"MICR Number",font=red_font,fill=grey_fill)
        write_cell(ws, cur_row, 4, 8,self.esic_bank_micr or "") 
        

        # add formating in all rows
        for row in range(1, cur_row + 1):
            for col_idx in range(1, 9):
                col = get_column_letter(col_idx)
                cell = ws[f"{col}{row}"]

                if cell.font == Font():
                    cell.font = font_all
                if cell.border == white_border:
                    continue
                else:
                    cell.border = border
                if cell.alignment == Alignment():
                    cell.alignment = align_left
                    
        # ============================= Extra Details ====================================
        
        new_cur_row = 2
        write_cell(ws, new_cur_row, 10, 20,"Note: Red Mark Details Compulsory Fill",font=red_font,border=white_border)
        
        new_cur_row += 1
        write_cell(ws, new_cur_row, 10, 20,"OLD ESIC NO",font=red_font,border=white_border)
        
        new_cur_row +=1
        write_cell(ws, new_cur_row, 10, 20,"",border=white_border)
        
        new_cur_row += 1
        write_cell(ws, new_cur_row, 10, 20,"Compulsorily Required Documents.",font=red_font,border=white_border)
        
        new_cur_row += 1
        write_cell(ws, new_cur_row, 10, 20,"1). Canceled/ Bank passbook first page with JPG format",font=red_font,border=white_border)
        
        new_cur_row += 1
        write_cell(ws, new_cur_row, 10, 20,"2). Aadhar Card",font=red_font,border=white_border)
        
        new_cur_row += 1
        write_cell(ws, new_cur_row, 10, 20,"3). Passport size photo of Insrured person/employee is compulsorily required.",font=red_font,border=white_border)
        
        new_cur_row += 1
        write_cell(ws, new_cur_row, 10, 20,"4) if employee is giving the family details, passport size photo required",font=red_font,border=white_border)

        
        # Save workbook
        wb.save(output)
        output.seek(0)

        xls_file = base64.b64encode(output.read())
        output.close()

        # Create attachment
        filename = (
            f"ESIC - {self.employee_id.name or ''}({self.employee_id.employee_code or ''}).xls"
            if self.employee_id.employee_code
            else f"ESIC - {self.employee_id.name or ''}.xls"
        )
        
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': xls_file,
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        # Return download action
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }
