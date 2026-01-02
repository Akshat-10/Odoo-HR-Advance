from odoo import models, fields
from odoo.exceptions import UserError
from io import BytesIO
import base64
from openpyxl import Workbook


class NetSalaryWizard(models.TransientModel):
    _name = "net.salary.wizard"
    _description = "Net Salary Wizard"

    date_from = fields.Date(string="From Date", required=True)
    date_to = fields.Date(string="To Date", required=True)

    # 🔴 ADDED: Checkbox for all employees
    all_employee = fields.Boolean(string="All Employees")

    # Existing field (unchanged)
    employee_ids = fields.Many2many(
        'hr.employee',
        string="Employees"
    )

    def action_generate_excel(self):
        if self.date_from > self.date_to:
            raise UserError("From Date cannot be greater than To Date")

        import io
        import base64
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.drawing.image import Image
        from PIL import Image as PILImage
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Net Salary"

        # ================= STYLES =================
        border = Border(
            top=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin'),
            bottom=Side(style='thin')
        )
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

        font_header = Font(name='Arial', size=12, bold=True)
        font_regular = Font(name='Arial', size=11)

        blue_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
        light_blue_fill = PatternFill(start_color='87CEFA', end_color='87CEFA', fill_type='solid')
        grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

        ws.row_dimensions[1].height = 60
        ws.row_dimensions[2].height = 20

        # ================= EMPLOYEES =================
        if self.all_employee:
            employees = self.env['hr.employee'].search([
                ('active', '=', True),
                ('company_id', '=', self.env.company.id),
            ])
        elif self.employee_ids:
            employees = self.employee_ids
        else:
            employees = self.env['hr.employee']

        # ================= FIXED BANK COLUMNS =================
        fixed_banks = ["KOTAK", "ICICI"]

        bank_names = []
        for emp in employees:
            if emp.bank_account_id and emp.bank_account_id.bank_id:
                bank = emp.bank_account_id.bank_id.name.strip().upper()
                if bank not in fixed_banks and bank not in bank_names:
                    bank_names.append(bank)

        header_row = 3
        col = 1

        ws.cell(header_row, col, "Emp Code")
        col += 1

        bank_col_map = {}

        for bank in fixed_banks:
            ws.cell(header_row, col, f"{bank} BANK AC NO")
            bank_col_map[bank] = col
            col += 1

        for bank in bank_names:
            ws.cell(header_row, col, f"{bank} BANK AC NO")
            bank_col_map[bank] = col
            col += 1

        base_headers = [
            "PF?", "ESIC?", "UAN", "ESIC NO", "Employee Name",
            "New Gross", "Department Name", "Designation Name",
            "CTC salary",

            "GROSS SALARY (Calc)", "Basic (Calc)", "HRA (Calc)",
            "Conveyance (Calc)", "LTA (Calc)", "Other Allowance (Calc)",
            "TOTAL (Calc)", "Bonus (Calc)", "Payable (Calc)",
            "PF SALARY (Calc)", "Paid Days (Calc)",

            "GROSS SALARY", "Basic", "HRA", "Conveyance", "LTA",
            "Other Allowance", "TOTAL", "Bonus", "Payable",
            "PF", "Paid SALARY", "Paid Days"
        ]

        base_col_start = col
        for h in base_headers:
            ws.cell(header_row, col, h)
            col += 1

        last_col = col - 1
        last_col_letter = get_column_letter(last_col)

        # ================= DYNAMIC COMPANY TITLE =================
        ws.merge_cells(f"A1:{last_col_letter}1")
        ws["A1"] = "GERMAN TMT PVT LTD"
        ws["A1"].font = Font(size=20, bold=True)
        ws["A1"].alignment = align_center
        ws["A1"].fill = light_blue_fill

        ws["A2"] = "Month:"
        ws["A2"].font = font_header
        ws["A2"].fill = grey_fill
        ws.merge_cells("B2:C2")
        ws["B2"] = self.date_from.strftime('%B %Y')

        ws["D2"] = "Days"
        ws["D2"].font = font_header
        ws["D2"].fill = grey_fill
        ws["E2"] = (self.date_to - self.date_from).days + 1

        for r in (1, 2):
            for c in range(1, last_col + 1):
                ws.cell(r, c).border = border

        for c in range(1, last_col + 1):
            cell = ws.cell(header_row, c)
            cell.fill = blue_fill
            cell.font = Font(bold=True)
            cell.alignment = align_center
            cell.border = border
            ws.column_dimensions[cell.column_letter].width = 18

        # ================= PAYSLIPS =================
        payslips = self.env['hr.payslip'].search([
            ('employee_id', 'in', employees.ids),
            ('date_from', '<=', self.date_to),
            ('date_to', '>=', self.date_from),
            ('state', 'in', ['done', 'paid']),
        ])
        payslip_map = {p.employee_id.id: p for p in payslips}

        row = header_row + 1

        for emp in employees:
            slip = payslip_map.get(emp.id)
            lines = {l.code: l.total for l in slip.line_ids} if slip else {}

            new_gross = 0.0
            bonus = lines.get('BONUS', 0.0)
            pf = lines.get('PF', 0.0)
            esic = lines.get('ESIC', 0.0)
            paid_days = sum(slip.worked_days_line_ids.mapped('number_of_days')) if slip else 0

            # ================= 🔴 FIX START =================
            # Employee Code (via related field from hr.employee)
            ws.cell(row, 1, emp.employee_code or '')
            # ================= 🔴 FIX END =================

            base_data = [
                "YES" if pf else "NO",
                "YES" if esic else "NO",
                emp.l10n_in_uan or '',
                emp.l10n_in_esic_number or '',
                emp.name,
                new_gross,
                emp.department_id.name if emp.department_id else '',
                emp.job_id.name if emp.job_id else '',
                emp.contract_id.final_yearly_costs if emp.contract_id else 0,

                0, 0, 0, 0, 0, 0, 0, bonus, 0, pf, paid_days,

                lines.get('GROSS', 0.0),
                lines.get('BASIC', 0.0),
                lines.get('HRA', 0.0),
                lines.get('CONV', 0.0),
                lines.get('LTA', 0.0),
                lines.get('ADHOC', 0.0),
                sum(lines.get(k, 0.0) for k in ['BASIC', 'HRA', 'CONV', 'LTA', 'ADHOC', 'BONUS']),
                bonus,
                lines.get('NET', 0.0),
                pf,
                lines.get('NET', 0.0),
                paid_days
            ]

            c = base_col_start
            for v in base_data:
                ws.cell(row, c, v)
                c += 1

                # 🔴 ADDED: Bank Account Number Fetch (WORKING CODE)
            if emp.bank_account_id and emp.bank_account_id.bank_id:
                bank = emp.bank_account_id.bank_id.name.strip().upper()
                if bank in bank_col_map:
                    ws.cell(row, bank_col_map[bank], emp.bank_account_id.acc_number or '')

            col_map = {h: get_column_letter(base_col_start + i) for i, h in enumerate(base_headers)}


            # ================= ONLY FORMULA FIX =================

            # ✅ GROSS SALARY (Calc) = New Gross - Bonus  (I - T)
            ws[f"{col_map['GROSS SALARY (Calc)']}{row}"] = (
                f"={col_map['New Gross']}{row}-{col_map['Bonus (Calc)']}{row}"
            )

            ws[f"{col_map['Basic (Calc)']}{row}"] = f"=ROUND({col_map['GROSS SALARY (Calc)']}{row}*0.5,0)"
            ws[f"{col_map['HRA (Calc)']}{row}"] = f"=ROUND({col_map['Basic (Calc)']}{row}*0.4,0)"
            ws[f"{col_map['Conveyance (Calc)']}{row}"] = f"=ROUND({col_map['Basic (Calc)']}{row}*0.1,0)"
            ws[f"{col_map['LTA (Calc)']}{row}"] = f"=ROUND({col_map['Basic (Calc)']}{row}*0.1,0)"

            # Other Allowance = M - N - O - P - Q (already correct)
            ws[f"{col_map['Other Allowance (Calc)']}{row}"] = (
                f"={col_map['GROSS SALARY (Calc)']}{row}"
                f"-{col_map['Basic (Calc)']}{row}"
                f"-{col_map['HRA (Calc)']}{row}"
                f"-{col_map['Conveyance (Calc)']}{row}"
                f"-{col_map['LTA (Calc)']}{row}"
            )

            ws[f"{col_map['TOTAL (Calc)']}{row}"] = (
                f"=SUM({col_map['Basic (Calc)']}{row}:{col_map['Other Allowance (Calc)']}{row})"
            )

            ws[f"{col_map['Payable (Calc)']}{row}"] = (
                f"={col_map['TOTAL (Calc)']}{row}+{col_map['Bonus (Calc)']}{row}"
            )

            ws[f"{col_map['PF SALARY (Calc)']}{row}"] = (
                f'=IF({col_map["PF?"]}{row}="NO",0,'
                f'IF({col_map["PF?"]}{row}="YES",'
                f'IF(({col_map["Basic (Calc)"]}{row}'
                f'+{col_map["Conveyance (Calc)"]}{row}'
                f'+{col_map["LTA (Calc)"]}{row}'
                f'+{col_map["Other Allowance (Calc)"]}{row})>=15000,15000,'
                f'ROUND(({col_map["Basic (Calc)"]}{row}'
                f'+{col_map["Conveyance (Calc)"]}{row}'
                f'+{col_map["LTA (Calc)"]}{row}'
                f'+{col_map["Other Allowance (Calc)"]}{row}),0))))'
            )

            ws[f"{col_map['Paid Days (Calc)']}{row}"] = f"={col_map['Paid Days']}{row}"

            # 🔴 ONLY ADDITION: DATA ROW BORDER (ROW 4 ONWARDS)
            for c in range(1, last_col + 1):
                ws.cell(row, c).border = border
                ws.cell(row, c).alignment = align_left
                ws.cell(row, c).font = font_regular

            row += 1

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        attachment = self.env['ir.attachment'].create({
            'name': 'Net_Salary_Report.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(output.read()),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }







